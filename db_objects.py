# -*- coding: utf-8 -*-

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.

# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from PyQt5.QtWidgets import QMessageBox, QFileDialog
from exceptions import *
from copy import copy
from openpyxl import Workbook


class Account():
  def __init__(self, account_id, name, brokerage, account_no):
    self.id = account_id
    self.name = name
    self.brokerage = brokerage
    self.account_no = account_no

    #dictionary to map fund ids to fund names
    self.fund_names = {}
    #dictionary to map AccountValue ids (account valuation on specific date) to AccountValue objects
    self.account_values_by_id = {}
    #dictionary to map fund.id to initial units
    self.initial_fund_units = {}
    #dictionary to map fund.id to ending units
    self.end_units = {} # This is initialized in processPurchases
    #account values list
    self.account_values_sorted_by_date = []

    # flag that the account has already been initialized
    self.initialized = False

  def __str__(self):
    return("Name = %s, Brokerage = %s, Account Number = %s" % (self.name, self.brokerage, self.account_no))

  def copy(self, account):
    self.id = account.id
    self.name = account.name
    self.brokerage = account.brokerage
    self.account_no = account.account_no

  def insertIntoDB(self, con):
    cursor = con.execute("INSERT INTO Accounts (name, brokerage, account_no) VALUES (?, ?, ?)",
                         (self.name, self.brokerage, self.account_no))
    self.id = cursor.lastrowid
    con.commit()

  def fetchFunds(self, con):
    self.funds = []
    cursor = con.cursor()
    for row in cursor.execute("SELECT id, name, initial_units, account_id FROM Funds WHERE account_id = ?", (self.id,)):
      self.funds += [makeFund(row)]
      self.fund_names[row[0]] = row[1]

  def fundNameExists(self, fund_name):
    return fund_name in self.fund_names.values()

  def addFund(self, new_fund, con):
    self.funds += [new_fund]
    self.fund_names[new_fund.id] = new_fund.name
    self.initialUnitValues()
    self.processPurchases(con)

  def fundChanged(self, con):
    self.initialUnitValues()
    self.processPurchases(con)

  def deleteFund(self, fund, con):
    fund.deletePurchases(con)
    con.execute("DELETE FROM Funds WHERE id = ?", (fund.id,))
    con.commit()
    del self.fund_names[fund.id]
    del self.initial_fund_units[fund.id]
    self.funds.remove(fund)
    self.processPurchases(con)

  def deleteAllFunds(self, con):
    con.execute("DELETE FROM Funds WHERE account_id = ?", (self.id,))
    con.commit()

  def deleteAllAccountValues(self, con):
    con.execute("DELETE FROM AccountValue WHERE account_id = ?", (self.id,))
    con.commit()

  def deleteAccount(self, con):
    self.deleteAllFunds(con)
    self.deleteAllAccountValues(con)
    con.execute("DELETE FROM Accounts WHERE id = ?", (self.id,))
    con.commit()

  def updateToDB(self, con):
    con.execute("UPDATE Accounts SET name = ?, brokerage = ?, account_no = ? WHERE id = ?",
                (self.name, self.brokerage, self.account_no, self.id))
    con.commit()

  def fetchValues(self, con):
    cursor = con.cursor()
    for row in cursor.execute("SELECT id, date, value, account_id FROM AccountValue WHERE account_id = ? ORDER BY date", (self.id,)):
      v = makeAccountValue(row)
      self.account_values_sorted_by_date += [v]

  def addValue(self, value):
    self.account_values_sorted_by_date += [value]
    self.account_values_sorted_by_date.sort(key = lambda av: av.date)

  def initialize(self, con):
    if (not self.initialized):
      self.fetchFunds(con)
      self.fetchValues(con)
      self.initialUnitValues()
      self.processPurchases(con)
      self.initialized = True

  def initialUnitValues(self):
    for f in self.funds:
      self.initial_fund_units[f.id] = f.initial_units

  def initialUnitValuesIsZero(self):
    return sum(self.initial_fund_units.values()) == 0

  def processPurchases(self, con):
    self.purchases = []
    last_units = copy(self.initial_fund_units)
    if not self.initialUnitValuesIsZero():
      for v in self.account_values_sorted_by_date:
        self.account_values_by_id[v.id] = v
        total_units = sum(last_units.values())
        v.unit_price = v.value/total_units
        purchases = v.fetchUnitPurchases(con, self.funds)
        for p in purchases:
          p.units_purchased = p.amount/v.unit_price
          last_units[p.fund_id] += p.units_purchased
          self.purchases += [p]
        v.units_out = copy(last_units)
    self.end_units = copy(last_units)
    self.total_units = sum(self.end_units.values())

  def save_workbook_with_retry(self, wbk, filename, parent=None):
      while True:
          try:
              wbk.save(filename)
              QMessageBox.information(parent, "Success", f"File saved successfully:\n{filename}")
              break
          except PermissionError:
              retry = QMessageBox.warning(
                  parent,
                  "File is Open",
                  f"The file:\n{filename}\nis open in Excel or another program.\n"
                  "Please close it or choose a new filename.",
                  QMessageBox.Retry | QMessageBox.Cancel
              )
              if retry == QMessageBox.Cancel:
                  break
              filename, _ = QFileDialog.getSaveFileName(parent, "Save As", filename, "Excel Files (*.xlsx)")
              if not filename:
                  break
          except Exception as e:
              QMessageBox.critical(parent, "Error", f"An unexpected error occurred:\n{str(e)}")
              break

  def exportXLSX(self, file_name):
    wbk = Workbook()
    funds_sheet = wbk.active
    funds_sheet.title = "Funds"
    funds_sheet.append(["Name", "Initial Units", "Ending Units", "% of Total"])
    funds_sheet.column_dimensions['A'].width = 25
    funds_sheet.column_dimensions['B'].width = 15
    funds_sheet.column_dimensions['C'].width = 15
    funds_sheet.column_dimensions['D'].width = 15
    for f in self.funds:
      funds_sheet.append([f.name, f.initial_units, self.end_units[f.id],
                          (self.end_units[f.id] / self.total_units * 100)])
    funds_sheet.append(["", "Total", self.total_units, ""])
    purchases_sheet = wbk.create_sheet("Purchases")
    purchases_sheet.append(["Date", "Fund Name", "Amount", "Units Purchased"])
    purchases_sheet.column_dimensions['A'].width = 15
    purchases_sheet.column_dimensions['B'].width = 25
    purchases_sheet.column_dimensions['C'].width = 15
    purchases_sheet.column_dimensions['D'].width = 15
    for p in self.purchases:
      purchases_sheet.append([self.account_values_by_id[p.date_id].date,
                              self.fund_names[p.fund_id],
                              p.amount, p.units_purchased])
    account_values_sheet = wbk.create_sheet("Account Values")
    account_values_sheet.append(["Date", "Account Value"])
    account_values_sheet.column_dimensions['A'].width = 15
    account_values_sheet.column_dimensions['B'].width = 15
    for av in self.account_values_sorted_by_date:
      account_values_sheet.append([av.date, av.value])
    self.save_workbook_with_retry(wbk, file_name, parent=None)


def makeAccount(tuple4):
  if len(tuple4) != 4:
    raise TupleLengthError()
  return Account(tuple4[0], tuple4[1], tuple4[2], tuple4[3])


class Fund():
  def __init__(self, fund_id, name, initial_units,  account_id):
    self.id = fund_id
    self.name = name
    self.initial_units= initial_units
    self.account_id = account_id

  def __str__(self):
    return "id = %d, name = %s, initial_units = %f, account_id = %d" \
      % (self.id, self.name, self.initial_units, self.account_id)

  def copy(self, fund):
    self.id = fund.id
    self.name = fund.name
    self.initial_units = fund.initial_units
    self.account_id = fund.account_id

  def insertIntoDB(self, con):
    cursor = con.execute("INSERT INTO Funds (name, initial_units, account_id) VALUES (?, ?, ?)",
                         (self.name, self.initial_units, self.account_id))
    self.id = cursor.lastrowid
    con.commit()

  def updateToDB(self, con):
    con.execute("UPDATE Funds SET name = ?, initial_units = ?, account_id = ? WHERE id = ?",
                (self.name, self.initial_units, self.account_id, self.id))
    con.commit()

  def deletePurchases(self, con):
    con.execute("DELETE FROM UnitPurchase WHERE fund_id = ?", (self.id,))
    con.commit()


def makeFund(tuple4):
  if len(tuple4) != 4:
    raise TupleLengthError()
  return Fund(tuple4[0], tuple4[1], tuple4[2], tuple4[3])


class AccountValue():
  def __init__(self, account_value_id, date, value, account_id):
    self.id = account_value_id
    self.date = date
    self.value = value
    self.account_id = account_id

  def __str__(self):
    return "id = %d, date = %s, value = %f, account_id = %d" % (self.id, self.date, self.value, self.account_id)

  def copy(self, account_value):
    self.id = account_value.id
    self.date = account_value.date
    self.value = account_value.value
    self.account_id = account_value.account_id

  def insertIntoDB(self, con):
    cursor = con.execute("INSERT INTO AccountValue (date, value, account_id) VALUES (?, ?, ?)",
                         (self.date, self.value, self.account_id))
    self.id = cursor.lastrowid
    con.commit()

  def updateToDB(self, con):
    con.execute("UPDATE AccountValue SET date = ?, value = ?, account_id = ? WHERE id = ?",
                (self.date, self.value, self.account_id, self.id))
    con.commit()

  def fetchUnitPurchases(self, con, funds):
    purchases = []
    cursor = con.cursor()
    funds_ids = tuple(f.id for f in funds)
    query = "SELECT UnitPurchase.id, fund_id, amount, date_id FROM UnitPurchase JOIN AccountValue ON UnitPurchase.date_id = AccountValue.id WHERE fund_id IN ({seq}) AND date_id = ?".format(
        seq=','.join(['?']*len(funds_ids)))
    params = list(funds_ids) + [self.id]
    for row in cursor.execute(query, params):
      purchases += [makeUnitPurchase(row)]
    return purchases


def makeAccountValue(tuple4):
  if len(tuple4) != 4:
    raise TupleLengthError()
  return AccountValue(tuple4[0], tuple4[1], tuple4[2], tuple4[3])


class UnitPurchase():
  def __init__(self, purchase_id, fund_id, amount, date_id):
    self.id = purchase_id
    self.fund_id = fund_id
    self.amount = amount
    self.date_id = date_id

  def __str__(self):
    return "id = %d, fund_id = %d, amount = %f, date_id = %d" % (self.id, self.fund_id, self.amount, self.date_id)

  def copy(self, unit_purchase):
    self.id = unit_purchase.id
    self.fund_id = unit_purchase.fund_id
    self.amount = unit_purchase.amount
    self.date_id = unit_purchase.date_id

  def insertIntoDB(self, con):
    cursor = con.execute("INSERT INTO UnitPurchase (fund_id, amount, date_id) VALUES (?, ?, ?)",
                         (self.fund_id, self.amount, self.date_id))
    self.id = cursor.lastrowid
    con.commit()

  def updateToDB(self, con):
    con.execute("UPDATE UnitPurchase SET fund_id = ?, amount = ?, date_id = ? WHERE id = ?",
                (self.fund_id, self.amount, self.date_id, self.id))
    con.commit()

  def deleteFromDB(self, con):
    con.execute("DELETE FROM UnitPurchase WHERE id = ?", (self.id,))
    con.commit()


def makeUnitPurchase(tuple4):
  if len(tuple4) != 4:
    raise TupleLengthError()
  return UnitPurchase(tuple4[0], tuple4[1], tuple4[2], tuple4[3])


def fillCell(s, r, c, v):
  s.cell(row = r, column = c).value = v
